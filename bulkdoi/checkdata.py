import openpyxl
import sys
import re

def extract_header(datafile):
    wb = openpyxl.load_workbook(filename=datafile)
    sh = wb.active
    header = []
    for i in range(1, sh.max_column+1):
        cell_obj = sh.cell(row=1, column=i)
        if cell_obj.value is None: break
        header.append(cell_obj.value.strip())
    return header

def extract_data(datafile):
    revised_header = ['url', 'creators', 'title', 'publisher', 'pubyear', 'restype', 'description']
    wb = openpyxl.load_workbook(filename=datafile)
    sh = wb.active
    data = []
    numcols = len(revised_header)
    for row in sh.iter_rows(min_row=2):
        rowdata = []
        if row[0].value:
            for idx in range(numcols):
                cellval = row[idx].value
                if cellval is None:
                    rowdata.append('')
                else:
                    rowdata.append(row[idx].value)
        if rowdata:
            data.append(dict(zip(revised_header, rowdata)))
    return data



def checkheader(header):
    if header_is_malformed(header):
        return (False, 'Data file header is malformed')
    return (True, '')

def checkdata(data):
    errors = []
    for idx, row in enumerate(data):
        rowerrors = checkrow(row)
        if rowerrors:
            errors.append((idx+2, rowerrors))
    return errors

def header_is_malformed(header):
    expected_header = ['URL', 'Creators', 'Title', 'Publisher', 'Publication Year', 'Resource Type', 'Description']
    if len(header) != len(expected_header):
        return True
    for i, j in zip(header, expected_header):
        if i.lower() != j.lower():
            return True
    return False

def checkrow(row):
    errors = []
    errors.append(url_is_malformed(row['url']))
    errors.append(creators_is_malformed(row['creators']))
    errors.append(title_is_malformed(row['title']))
    errors.append(publisher_is_malformed(row['publisher']))
    errors.append(pubyear_is_malformed(row['pubyear']))
    errors.append(restype_is_malformed(row['restype']))
    return [i for i in errors if i]

def url_is_malformed(url):
    expected_form = url.startswith('https://') or url.startswith('http://') or url.startswith('ftp://')
    if expected_form:
        return None
    return 'URL field is malformed'

def creators_is_malformed(creators):
    if not len(creators.strip()):
        return 'Creators field is empty'
    for item in creators.split(';'):
        if item.startswith('['):
            if not re.fullmatch('^\[[^\[\]]+\]$', item.strip()):
                return 'Creators field has an organization with mismatched brackets'
        else:
            if '[' in item:
                return 'Creators field has a name with an embedded ['
            if ']' in item:
                return 'Creators field has a name with an embedded ]'
            name = [n.strip() for n in item.split(',')]
            if len(name) == 0 or len(name) > 2:
                return 'Creators field has a name with multiple commas'

def title_is_malformed(title):
    if len(title.strip()):
        return None
    return 'Title field is empty'

def publisher_is_malformed(publisher):
    if len(publisher.strip()):
        return None
    return 'Publisher field is empty'

def pubyear_is_malformed(pubyear):
    try:
        pubyear_int = int(pubyear)
    except:
        return 'Publication Year field must be an integer'
    if pubyear_int > 1900:
        return None
    return 'Publication Year field is malformed'

def restype_is_malformed(restype):
    accepted = set([
        'Audiovisual',
        'Collection',
        'DataPaper',
        'Dataset',
        'Event',
        'Image',
        'InteractiveResource',
        'Model',
        'PhysicalObject',
        'Service',
        'Software',
        'Sound',
        'Text',
        'Workflow'
    ])
    if restype.strip() in accepted:
        return None
    return 'Resource Type field must be one of the choices in list'

def main(argv=None):
    if argv is None:
        argv = sys.argv
    datafile = argv[1]
    header = extract_header(datafile)
    print(checkheader(header))
    data = extract_data(datafile)
    print(checkdata(data))


if __name__ == "__main__":
    sys.exit(main())