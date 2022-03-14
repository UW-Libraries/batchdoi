import sys
import openpyxl


def extract_header(datafile):
    wb = openpyxl.load_workbook(filename=datafile)
    sh = wb.active
    header = []
    for i in range(1, sh.max_column+1):
        cell_obj = sh.cell(row=1, column=i)
        if cell_obj.value is None: break
        header.append(cell_obj.value.strip())
    return header


def extract_data(datafile):
    revised_header = ['url', 'creators', 'title', 'publisher', 'publication_year', 'resource_type', 'description']
    wb = openpyxl.load_workbook(filename=datafile)
    sh = wb.active
    data = []
    numcols = len(revised_header)
    for row in sh.iter_rows(min_row=2):
        rowdata = []
        if row[0].value:
            for idx in range(numcols):
                cellval = row[idx].value
                if cellval is None:
                    rowdata.append('')
                else:
                    rowdata.append(row[idx].value)
        if rowdata:
            data.append(dict(zip(revised_header, rowdata)))
    return data


def main(argv=None):
    if argv is None:
        argv = sys.argv
    datafile = argv[1]
    header = extract_header(datafile)
    print(checkheader(header))
    data = extract_data(datafile)
    print(data)


if __name__ == "__main__":
    sys.exit(main())